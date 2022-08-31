import openpyxl
from openpyxl import Workbook

#  https://openpyxl.readthedocs.io/en/stable/tutorial.html#create-a-workbook

wb = Workbook()

ws = wb.active

#Para crear una nueva hoja de trabajo

ws1 = wb.create_sheet("Mysheet2", 0) # insert at first position

ws2 = wb.create_sheet("Mysheet3", -1) # insert at the penultimate position

ws3 = wb.create_sheet("Mysheet1") # insert at the end (default)


# Para cambiar el nombre de una hoja de trabajo

ws2.title = "Lista Procesada"

ws1.title = "Lista de Entrada"

# Para cambiar el color de la pestaña de la hoja de trabajo

ws2.sheet_properties.tabColor = "1072BA"

# Se puede obtener la hoja de trabajo a traves del nombre de la misma

ws3 = wb["Mysheet1"]

# Para obtener los valores de todas las hojas de trabajo

print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)

# Para crear copias de las hojas de trabajo dentro de un libro de trabajo
"""You cannot copy a worksheet if the workbook is open in read-only or write-only mode."""

source = wb.active
target = wb.copy_worksheet(source)

# Como acceder a una celda de una hoja de trabajo de una libro de trabajo
"""Se puede acceder directamente con la coordena de la celda"""

c = ws1['A4']

"""Para asignar un valor a una determinada celda"""

ws1['A4'] = 4

"""Podemos realizar lo mismo mediante el metodo"""
"""Cuando se crea una hoja de cálculo en la memoria, no contiene celdas. Se crean cuando se accede por primera vez."""

d = ws1.cell(row = 4, column = 2, value = "Parace que se puede")

"""Debido a esta función, desplazarse por las celdas en lugar de acceder a ellas directamente las creará todas
en la memoria, incluso si no les asigna un valor."""

for x in range(1,101):
    for y in range(1,101):
        ws1.cell(row = x, column = y, value = "German")

"""Podemos acceder a un rango de celdas"""

cell_range = ws1['A1':'C2']

"""Podemos obtener un rango de columnas o filas"""

colC = ws1['C']
col_range = ws1['C:D']
row10 = ws1[10]
row_range = ws1[5:10]

"""Podemos usar el siguiente metodo para acceder a las filas"""

for row in ws1.iter_rows(min_row = 1, max_col = 3, max_row = 2):
    for cell in row:
        print(cell)

"""Podemos usar el siguiente metodo para acceder a las columnas"""

for col in ws2.iter_cols(min_row = 1, max_col = 3, max_row = 2):
    for cell in col:
        print(cell)

"""Para iterar todas las columnas y filas podemos usar el siguiente metodo"""

ws2 = wb.active
ws['C9'] = 'hello world'
tuple(ws.rows)

tuple(ws2.columns)

""""Para obtener los valores de las celdas podemos usar la propiedad"""

for row in ws2.values:
   for value in row:
     print(value)


"""Para asiganar un valor a una celda"""

c.value = 'hello, world'
print(c.value)

d.value = 3.14
print(d.value)

"""Para guardar un documento. CUIDADO, reescribe el documento!!!"""

wb.save('prueba.xlsx')


"""Para abrir un documento existente"""

from openpyxl import load_workbook
wb2 = load_workbook('test.xlsx')
print(wb2.sheetnames)





"""Para guardar un documento"""

""""
excel_document = openpyxl.load_workbook("sample.xlsx")

excel_document.get_sheet_names()


sheet = excel_document.get_sheet_by_name('Sheet1')

print (sheet['A2'].value)"""